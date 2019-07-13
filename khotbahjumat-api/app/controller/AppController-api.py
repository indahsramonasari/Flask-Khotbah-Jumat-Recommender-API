from app import app
from openpyxl import load_workbook
from flask import request, jsonify
from app.constant import RequestMethod
from app.model.QueriesModel import Queries
from app.model.DetailsModel import Details
from app.module.Engine import Engine, preprocess
from flask import request, render_template
import os
import requests
import pandas as pd
from math import log

pd.set_option('display.max_colwidth', -1)

# Load dataset first when open app from browser
dataset = pd.read_excel("app/db/preprocessed-dataset.xlsx")


@app.route("/", methods=RequestMethod.GET)
def index():
    return render_template("index.html")

@app.route("/cariquery", methods=RequestMethod.GET)
def cariquery():
    response = list()  # Define response
    if request.method == "POST":
        if "files" in request.files:
            file = request.files["files"]
            file.save(os.path.join("app/db", "queries.xlsx"))
            queries = pd.read_excel("app/db/queries.xlsx")
            queries = queries["Queries"].values
        else:
            resp = {
                "error": "invalid request",
                "path": "/cariquery",
                "message": "request should be file"
            }
            resp = jsonify(resp)
            resp.status_code = 400
            print(resp)
            return resp

    elif request.method == "GET":
        if 'q' in request.args:
            queries = [request.args['q']]
        else:
            resp = {
                "error": "invalid request",
                "path": "/cariquery",
                "message": "request should be query"
            }
            resp = jsonify(resp)
            resp.status_code = 400
            return resp


    # Preproces queries
    queriesPre = list()
    for query in queries:
        queriesPre.append(preprocess(query))

    # Cek di database apakah ada data dengan query pada inputan ataupun file
    for query in queriesPre:
        data = Queries.findByQueryName(query)
        if data is not None:
            response.append(data)

    if len(response) is not 0:
        return jsonify(response)
    else:
        engine = Engine()
        docs = [str(x) for x in dataset['preprocessed']]
        documentsName = list()

        for i, doc in enumerate(docs):
            engine.addDocument(doc)
            documentsName.append("Document_{}".format(i + 1))

        for query in queriesPre:
            engine.setQuery(query)  # Set query pencarian

        titlesScores = engine.process_score()
        ScoreDf = (pd.DataFrame(titlesScores)).T
        ScoreDf.columns = queriesPre
        ScoreDf["Documents"] = documentsName
        ScoreDf["Link"] = dataset["Link"].values

        dfListed = list()
        for i in queriesPre:
            labels = list()
            for j in ScoreDf[i]:
                if j > 0.000:
                    labels.append(1)
                else:
                    labels.append(0)
            datadf = pd.DataFrame(ScoreDf[i])
            datadf["Documents"] = ScoreDf["Documents"]
            datadf["Labels"] = labels
            datadf["Link"] = ScoreDf["Link"].values
            dfListed.append(datadf.sort_values(by=[i], ascending=False))

        for i, df in enumerate(dfListed):
            dbQuery = Queries(queriesPre[i])
            for j in range(len(df["Documents"])):
                document = df["Documents"][j]
                label = int(df["Labels"][j])
                score = float(df[queriesPre[i]][j])
                link = df["Link"][j]
                data = document, label, score, link
                details = Details(data)
                dbQuery.details.append(details)
            dbQuery.save()

        for query in queriesPre:
            data = Queries.findByQueryName(query)
            response.append(data)

        return jsonify(response)

def getQueriesFromDatabase(queries):
    # Check on database where queries is exists
    details = list()
    for query in queries:
        data = Queries.findByQueryName(query)
        if data is not None:
            # Define list
            documents = list()
            links = list()
            score = list()
            for detail in data["details"]:
                # Get document and link from details
                documents.append(detail["document"])
                score.append(detail["score"])
                links.append(detail["link"])
            df = pd.DataFrame([documents, score, links]).T
            df.columns = ["Document", "Score", "Link"]
            details.append(df)
    return details

@app.route("/proses", methods=RequestMethod.GET_POST)
def proses():
    if request.method == "POST":
        if "files" in request.files:
            file = request.files["files"]
            file.save(os.path.join("app/db", "datasetbaru.xlsx"))
        else:
            resp = {
                "error": "invalid request",
                "path": "/proses",
                "message": "request should be file"
            }
        
    databr = pd.read_excel("app/db/datasetbaru.xlsx")
    preprocessed_br = []
    docs = databr.Dokumen.tolist()

    for doc in docs:
        preprocessed_br.append(preprocess(doc))

    databr['preprocessed'] = preprocessed_br

    databr.to_excel("app/db/datasetbaru.xlsx", engine="xlsxwriter", index=False)

        #datalama-panggil
    wb = load_workbook(filename = 'app/db/preprocessed-dataset.xlsx')
    sheet_ranges = wb['Sheet1']
    dat = pd.DataFrame(sheet_ranges.values)
    dat.columns = ['Dokumen','Link','preprocessed']
    data = dat[1:1000]

    asli = []
    for dt in data['Dokumen']:
        asli.append(dt)
    asli1 = []
    for dt1 in data['Link']:
        asli1.append(dt1)
    asli2 = []
    for dt2 in data['preprocessed']:
        asli2.append(dt2)


    workb = load_workbook(filename='app/db/datasetbaru.xlsx')
    sheet_ranges1 = workb['Sheet1']
    dat2 = pd.DataFrame(sheet_ranges1.values)
    dat2.columns = ['Dokumen','Link','preprocessed']
    data2 = dat2[1:500]
        
    rows = []
    for d in data2['Dokumen']:
        rows.append(d)
    rows1 = []
    for d1 in data2['Link']:
        rows1.append(d1)
    rows2 = []
    for d2 in data2['preprocessed']:
        rows2.append(d2)


    for row in rows:
        asli.append(row)
    for row1 in rows1:
        asli1.append(row1)
    for row2 in rows2:
        asli2.append(row2)

    datasetku = pd.read_excel("app/db/a.xlsx")
    datasetku['Dokumen'] = asli
    datasetku['Link'] = asli1
    datasetku['preprocessed'] = asli2

    datasetku.to_excel("app/db/preprocessed-dataset.xlsx", engine="xlsxwriter", index=False)
